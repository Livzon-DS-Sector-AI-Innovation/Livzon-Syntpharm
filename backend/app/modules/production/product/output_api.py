"""Product output API routes."""

import csv
import io
import uuid
from datetime import UTC, date, datetime, timedelta, timezone
from typing import Any

from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, not_, or_, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.deps import RequiredUser
from app.core.response import ApiResponse
from app.modules.production.product.models import Product
from app.modules.production.product.output_models import WORKSHOP_CHOICES, ProductOutput
from app.modules.production.product.output_schemas import (
    AnnualReviewResponse,
    ImportResponse,
    PreviewImportResponse,
    ProductOutputCreate,
    ProductOutputResponse,
    ProductOutputUpdate,
    UndoImportResponse,
)
from app.modules.production.product.output_service import ProductOutputService
from app.platform.integrations.feishu.bitable import BitableClient

router = APIRouter()


@router.get("/product-output/workshops", summary="获取车间列表")
async def get_workshops() -> Any:
    """获取所有车间列表"""
    return ApiResponse(data=WORKSHOP_CHOICES)


@router.get("/product-output", summary="获取产量记录列表")
async def get_product_outputs(
    current_user: RequiredUser,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=5000),
    workshop: str | None = None,
    product_id: uuid.UUID | None = None,
    product_name: str | None = None,
    batch_no: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    sort_by: str | None = Query(None, description="排序字段: batch_no, production_date, end_date, weight"),
    sort_order: str = Query("desc", description="排序方向: asc 或 desc"),
    db: AsyncSession = Depends(get_db),
) -> Any:
    """获取产量记录列表"""
    service = ProductOutputService(db)
    skip = (page - 1) * page_size
    records, total = await service.get_list(
        skip=skip,
        limit=page_size,
        workshop=workshop,
        product_id=product_id,
        product_name=product_name,
        batch_no=batch_no,
        start_date=start_date,
        end_date=end_date,
        sort_by=sort_by,
        sort_order=sort_order,
    )
    return ApiResponse(
        data=[ProductOutputResponse.model_validate(r) for r in records],
        meta={"page": page, "page_size": page_size, "total": total},
    )


@router.get("/product-output/summary", summary="获取汇总统计")
async def get_product_outputs_summary(
    current_user: RequiredUser,
    target_date: date | None = Query(None, description="查询日期"),
    month: str | None = Query(None, description="查询月份 YYYY-MM"),
    year: int | None = Query(None, description="查询年份"),
    product_id: uuid.UUID | None = Query(None, description="产品ID"),
    start_date: date | None = Query(None, description="开始日期"),
    end_date: date | None = Query(None, description="结束日期"),
    db: AsyncSession = Depends(get_db),
) -> Any:
    """获取每日/月/年汇总统计"""
    service = ProductOutputService(db)
    summary = await service.get_summary(
        target_date=target_date,
        month=month,
        year=year,
        product_id=product_id,
        start_date=start_date,
        end_date=end_date,
    )
    return ApiResponse(data=summary)


@router.get("/product-output/batch-count", summary="获取批次统计")
async def get_product_outputs_batch_count(
    current_user: RequiredUser,
    target_date: date | None = Query(None, description="查询日期"),
    month: str | None = Query(None, description="查询月份 YYYY-MM"),
    year: int | None = Query(None, description="查询年份"),
    product_id: uuid.UUID | None = Query(None, description="产品ID"),
    start_date: date | None = Query(None, description="开始日期"),
    end_date: date | None = Query(None, description="结束日期"),
    db: AsyncSession = Depends(get_db),
) -> Any:
    """获取批次统计"""
    service = ProductOutputService(db)
    batch_counts = await service.get_batch_count(
        target_date=target_date,
        month=month,
        year=year,
        product_id=product_id,
        start_date=start_date,
        end_date=end_date,
    )
    return ApiResponse(data=batch_counts)


@router.get("/product-output/export", summary="导出产量记录")
async def export_product_outputs(
    current_user: RequiredUser,
    workshop: str | None = None,
    product_id: uuid.UUID | None = None,
    product_name: str | None = None,
    batch_no: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    db: AsyncSession = Depends(get_db),
) -> Any:
    """导出产量记录为 XLSX"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    service = ProductOutputService(db)
    records, _ = await service.get_list(
        skip=0,
        limit=10000,
        workshop=workshop,
        product_id=product_id,
        product_name=product_name,
        batch_no=batch_no,
        start_date=start_date,
        end_date=end_date,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "产量记录"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["车间", "产品名称", "批号", "生产日期", "结束日期", "重量", "单位", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, r in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=r.workshop)
        ws.cell(row=row_idx, column=2, value=r.product_name)
        ws.cell(row=row_idx, column=3, value=r.batch_no)
        ws.cell(row=row_idx, column=4, value=r.production_date.isoformat() if r.production_date else "")
        ws.cell(row=row_idx, column=5, value=r.end_date.isoformat() if r.end_date else "")
        ws.cell(row=row_idx, column=6, value=r.weight)
        ws.cell(row=row_idx, column=7, value=r.unit)
        ws.cell(row=row_idx, column=8, value=r.notes or "")
        for col in range(1, 9):
            ws.cell(row=row_idx, column=col).border = thin_border

    column_widths = [12, 20, 15, 12, 12, 10, 8, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=product_outputs.xlsx"},
    )


@router.get("/product-output/annual-review", summary="获取年度回顾数据")
async def get_annual_review(
    current_user: RequiredUser,
    year: int = Query(..., description="年份"),
    db: AsyncSession = Depends(get_db),
) -> AnnualReviewResponse:
    """获取年度回顾数据"""
    service = ProductOutputService(db)
    return await service.get_annual_review(year)


@router.get("/product-output/annual-review/export", summary="导出年度回顾Excel")
async def export_annual_review(
    current_user: RequiredUser,
    year: int = Query(..., description="年份"),
    db: AsyncSession = Depends(get_db),
) -> Any:
    """导出年度回顾Excel"""
    from io import BytesIO

    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook

    service = ProductOutputService(db)
    data = await service.get_annual_review(year)

    wb = Workbook()

    # Sheet 1: 年度概览
    ws1 = wb.active
    ws1.title = "年度概览"
    ws1.append(["指标", "数值"])
    ws1.append(["年度总产量(kg)", data.overview.total_weight])
    ws1.append(["去年总产量(kg)", data.overview.previous_year_weight])
    ws1.append(["产量同比(%)", data.overview.weight_yoy])
    ws1.append(["年度总批次", data.overview.total_batches])
    ws1.append(["去年总批次", data.overview.previous_year_batches])
    ws1.append(["批次同比(%)", data.overview.batch_yoy])
    ws1.append(["活跃车间数", data.overview.active_workshops])
    ws1.append(["活跃产品数", data.overview.active_products])

    # Sheet 2: 月度趋势
    ws2 = wb.create_sheet("月度趋势")
    ws2.append(["月份", "当年产量(kg)", "去年产量(kg)"])
    for item in data.monthly_trend:
        ws2.append([item.month, item.current_year_weight, item.previous_year_weight])

    # Sheet 3: 车间排名
    ws3 = wb.create_sheet("车间排名")
    ws3.append(["排名", "车间", "总产量(kg)", "批次数"])
    for i, wr in enumerate(data.workshop_ranking, 1):
        ws3.append([i, wr.workshop, wr.total_weight, wr.batch_count])

    # Sheet 4: TOP产品
    ws4 = wb.create_sheet("TOP产品")
    ws4.append(["排名", "产品名称", "车间", "总产量(kg)", "批次数", "平均批次重量(kg)"])
    for tp in data.top_products:
        ws4.append([tp.rank, tp.product_name, tp.workshop, tp.total_weight, tp.batch_count, tp.avg_weight])

    # 输出Excel
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''annual_review_{year}.xlsx"},
    )


@router.get("/product-output/{record_id}", summary="获取产量记录详情")
async def get_product_output(
    record_id: uuid.UUID,
    current_user: RequiredUser,
    db: AsyncSession = Depends(get_db),
) -> Any:
    """获取单条产量记录"""
    service = ProductOutputService(db)
    record = await service.get_by_id(record_id)
    if not record:
        return ApiResponse(code=404, message="记录不存在")
    return ApiResponse(data=ProductOutputResponse.model_validate(record))


@router.post("/product-output", summary="新建产量记录")
async def create_product_output(
    data: ProductOutputCreate,
    current_user: RequiredUser,
    db: AsyncSession = Depends(get_db),
) -> Any:
    """新建产量记录"""
    service = ProductOutputService(db)
    record = await service.create(data)
    return ApiResponse(
        data=ProductOutputResponse.model_validate(record),
        message="创建成功",
    )


@router.put("/product-output/{record_id}", summary="更新产量记录")
async def update_product_output(
    record_id: uuid.UUID,
    data: ProductOutputUpdate,
    current_user: RequiredUser,
    db: AsyncSession = Depends(get_db),
) -> Any:
    """更新产量记录"""
    service = ProductOutputService(db)
    existing = await service.get_by_id(record_id)
    if not existing:
        return ApiResponse(code=404, message="记录不存在")
    record = await service.update(record_id, data)
    return ApiResponse(
        data=ProductOutputResponse.model_validate(record),
        message="更新成功",
    )


@router.delete("/product-output/batch", summary="批量删除产量记录")
async def batch_delete_product_outputs(
    current_user: RequiredUser,
    ids: str = Query(..., description="逗号分隔的记录 ID 列表"),
    db: AsyncSession = Depends(get_db),
) -> Any:
    """批量软删除产量记录"""
    id_list = [uuid.UUID(id_str.strip()) for id_str in ids.split(",") if id_str.strip()]
    if not id_list:
        return ApiResponse(code=400, message="未提供有效的记录 ID")
    service = ProductOutputService(db)
    count = 0
    for record_id in id_list:
        success = await service.delete(record_id)
        if success:
            count += 1
    return ApiResponse(data={"deleted": count}, message=f"成功删除 {count} 条记录")


@router.delete("/product-output/{record_id}", summary="删除产量记录")
async def delete_product_output(
    record_id: uuid.UUID,
    current_user: RequiredUser,
    db: AsyncSession = Depends(get_db),
) -> Any:
    """软删除产量记录"""
    service = ProductOutputService(db)
    success = await service.delete(record_id)
    if not success:
        return ApiResponse(code=404, message="记录不存在")
    return ApiResponse(message="删除成功")


async def parse_import_file(file: UploadFile, db: AsyncSession) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """解析导入文件，返回 (有效记录，无效记录)

    返回格式:
    - 有效记录：包含所有字段 + product_found (bool) + is_duplicate (bool)
    - 无效记录：包含错误信息
    """
    products_result = await db.execute(select(Product).where(not_(Product.is_deleted)))
    all_products = products_result.scalars().all()
    product_map: dict[tuple[str, str], uuid.UUID] = {}
    for p in all_products:
        workshop_key = p.workshop.replace(" ", "")
        product_map[(workshop_key, p.name)] = p.id

    filename = file.filename or ""
    file_ext = filename.lower().split(".")[-1] if "." in filename else ""
    content = await file.read()

    records_data = []
    invalid_records = []

    def parse_row(row_dict: dict[str, Any], row_num: int) -> None:
        workshop = str(row_dict.get("车间", "") or "").strip().replace(" ", "")
        product_name = str(row_dict.get("产品名称", "") or "").strip()
        batch_no = str(row_dict.get("批号", "") or "").strip()

        production_date_val = row_dict.get("生产日期")
        if production_date_val is not None and hasattr(production_date_val, "strftime"):
            production_date_str = production_date_val.strftime("%Y-%m-%d")
        else:
            production_date_str = str(production_date_val or "").strip()

        end_date_val = row_dict.get("结束日期")
        if end_date_val is not None and hasattr(end_date_val, "strftime"):
            end_date_str = end_date_val.strftime("%Y-%m-%d")
        else:
            end_date_str = str(end_date_val or "").strip()

        weight_str = str(row_dict.get("重量", "0") or "0").strip()
        unit = str(row_dict.get("单位", "kg") or "kg").strip() or "kg"
        notes = str(row_dict.get("备注", "") or "").strip() or None

        # 检查必填字段
        if not workshop or not product_name or not batch_no or not production_date_str:
            invalid_records.append({"row": row_num, "error": "缺少必填字段", "data": row_dict})
            return

        # 检查重量
        try:
            weight = float(weight_str)
        except ValueError:
            invalid_records.append({"row": row_num, "error": f"重量格式错误：{weight_str}", "data": row_dict})
            return

        # 匹配产品
        product_id = product_map.get((workshop, product_name))
        product_found = product_id is not None

        # 解析日期
        try:
            prod_date = date.fromisoformat(production_date_str)
            end_date = date.fromisoformat(end_date_str) if end_date_str else None
        except ValueError as e:
            invalid_records.append({"row": row_num, "error": f"日期格式错误：{e}", "data": row_dict})
            return

        records_data.append(
            {
                "workshop": workshop,
                "product_name": product_name,
                "product_id": product_id,
                "batch_no": batch_no,
                "production_date": prod_date,
                "end_date": end_date,
                "weight": weight,
                "unit": unit,
                "notes": notes,
                "product_found": product_found,
                "row_num": row_num,
            }
        )

    if file_ext == "xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(filename=io.BytesIO(content))
        ws = wb.active
        headers_row = [cell.value for cell in ws[1]]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            row_dict = dict(zip(headers_row, row))
            parse_row(row_dict, i)
    else:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for i, row in enumerate(reader, start=2):
            parse_row(row, i)

    return records_data, invalid_records


@router.post("/product-output/import/preview", summary="预览导入文件")
async def preview_import(
    current_user: RequiredUser,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
) -> Any:
    """预览导入文件，返回解析结果（不写入数据库）"""
    records_data, invalid_records = await parse_import_file(file, db)

    # 检查重复
    existing_result = await db.execute(
        select(ProductOutput.batch_no, ProductOutput.product_id).where(
            not_(ProductOutput.is_deleted),
        )
    )
    existing_pairs = set((row.batch_no, row.product_id) for row in existing_result.all())

    for record in records_data:
        record["is_duplicate"] = (record["batch_no"], record["product_id"]) in existing_pairs

    new_count = sum(1 for r in records_data if r["product_found"] and not r["is_duplicate"])
    duplicate_count = sum(1 for r in records_data if r["is_duplicate"])
    not_found_count = sum(1 for r in records_data if not r["product_found"])

    return ApiResponse(
        data=PreviewImportResponse(
            total_rows=len(records_data) + len(invalid_records),
            valid_records=len(records_data),
            invalid_records=len(invalid_records),
            new_records=new_count,
            duplicate_records=duplicate_count,
            not_found_product=not_found_count,
            records=records_data,
            invalid_details=invalid_records,
        )
    )


@router.post("/product-output/import", summary="导入产量记录")
async def import_product_outputs(
    current_user: RequiredUser,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
) -> Any:
    """通过 CSV 或 XLSX 文件批量导入产量记录"""
    import time

    batch_id = f"batch_{int(time.time())}"

    records_data, invalid_records = await parse_import_file(file, db)

    if not records_data:
        return ApiResponse(code=400, message="未找到有效数据，请检查文件格式")

    # 按 (批号 + 产品ID) 组合检查重复
    existing_result = await db.execute(
        select(ProductOutput.batch_no, ProductOutput.product_id).where(
            not_(ProductOutput.is_deleted),
        )
    )
    existing_pairs = set((row.batch_no, row.product_id) for row in existing_result.all())

    new_records = [r for r in records_data if (r["batch_no"], r["product_id"]) not in existing_pairs]
    skipped_count = len(records_data) - len(new_records)

    if not new_records:
        return ApiResponse(code=400, message=f"所有 {skipped_count} 条记录的批号已存在，跳过导入")

    # 添加批次ID
    for record in new_records:
        record["import_batch_id"] = batch_id
        # 移除预览用的临时字段
        record.pop("product_found", None)
        record.pop("row_num", None)

    service = ProductOutputService(db)
    count = await service.batch_import(new_records)

    message = f"成功导入 {count} 条记录"
    if skipped_count > 0:
        message += f"，跳过 {skipped_count} 条重复批号"

    return ApiResponse(
        data=ImportResponse(imported=count, skipped=skipped_count, batch_id=batch_id),
        message=message,
    )


@router.delete("/product-output/import/{batch_id}", summary="撤销导入批次")
async def undo_import(
    current_user: RequiredUser,
    batch_id: str,
    db: AsyncSession = Depends(get_db),
) -> Any:
    """撤销指定批次的导入记录"""
    result = await db.execute(
        select(ProductOutput).where(
            ProductOutput.import_batch_id == batch_id,
            not_(ProductOutput.is_deleted),
        )
    )
    records = result.scalars().all()

    if not records:
        return ApiResponse(code=404, message=f"未找到批次 {batch_id} 的记录")

    count = len(records)
    for record in records:
        record.is_deleted = True

    await db.commit()

    return ApiResponse(data=UndoImportResponse(deleted=count, batch_id=batch_id))


@router.post("/product-output/import-from-bitable", summary="从飞书多维表格导入产量记录")
async def import_from_bitable(
    current_user: RequiredUser,
    app_token: str = Query(..., description="飞书多维表格 app_token"),
    table_id: str = Query("", description="飞书多维表格 table_id（不填则自动获取第一个表）"),
    db: AsyncSession = Depends(get_db),
) -> Any:
    """从飞书多维表格批量导入产量记录"""
    from app.modules.production.product.feishu.bitable import ProductBitableClient

    if not app_token:
        return ApiResponse(code=400, message="请提供 app_token")

    resolved_table_id = table_id
    if not resolved_table_id:
        try:
            bitable_temp = BitableClient()
            bitable_temp.app_token = app_token
            tables = await bitable_temp.get_tables()
            if not tables:
                return ApiResponse(code=400, message="该多维表格中没有数据表")
            resolved_table_id = tables[0].get("table_id", "")
            if not resolved_table_id:
                return ApiResponse(code=400, message="无法自动获取表格 ID，请在链接中指定 table 参数")
        except Exception as e:
            return ApiResponse(code=500, message=f"获取表格列表失败：{str(e)}")

    try:
        bitable = ProductBitableClient(app_token=app_token, table_id=resolved_table_id)
        records = await bitable.search_records(page_size=500)

        if not records:
            return ApiResponse(code=400, message="多维表格中没有数据")

        def _extract_text(value: Any) -> str:
            if value is None:
                return ""
            if isinstance(value, list):
                if value and isinstance(value[0], dict):
                    return str(value[0].get("text", ""))
                return ""
            if isinstance(value, dict):
                return str(value.get("text", value.get("link", "")))
            return str(value)

        def _extract_number(value: Any) -> float:
            if value is None:
                return 0.0
            if isinstance(value, (int, float)):
                return float(value)
            if isinstance(value, list) and value:
                return float(value[0]) if value[0] else 0.0
            try:
                return float(str(value))
            except (ValueError, TypeError):
                return 0.0

        def _extract_date(value: Any) -> str:
            if value is None:
                return ""
            if isinstance(value, (int, float)):
                beijing_tz = timezone(timedelta(hours=8))
                dt = datetime.fromtimestamp(value / 1000, tz=UTC).astimezone(beijing_tz)
                return dt.strftime("%Y-%m-%d")
            if isinstance(value, list) and value:
                return _extract_date(value[0])
            return str(value)

        field_mapping = {
            "车间": "workshop",
            "生产车间": "workshop",
            "车间（生产车间）": "workshop",
            "产品名称": "product_name",
            "产品": "product_name",
            "产品名称（产品）": "product_name",
            "批号": "batch_no",
            "批次编号": "batch_no",
            "批号（批次编号）": "batch_no",
            "生产日期": "production_date",
            "生产开始日": "production_date",
            "生产日期（生产开始日）": "production_date",
            "结束日期": "end_date",
            "完工日": "end_date",
            "结束日期（完工日）": "end_date",
            "重量": "weight",
            "产出量": "weight",
            "重量（产出量）": "weight",
            "单位": "unit",
            "计量单位": "unit",
            "单位（计量单位）": "unit",
            "备注": "notes",
            "说明": "notes",
            "备注（说明）": "notes",
        }

        text_fields = {"workshop", "product_name", "batch_no", "unit", "notes"}
        number_fields = {"weight"}
        date_fields = {"production_date", "end_date"}

        records_data: list[dict[str, Any]] = []
        for record in records:
            fields = record.get("fields", {})
            if not fields:
                continue

            mapped: dict[str, Any] = {}
            for feishu_field, system_field in field_mapping.items():
                if feishu_field in fields:
                    raw_value = fields[feishu_field]
                    if system_field in text_fields:
                        mapped[system_field] = _extract_text(raw_value)
                    elif system_field in number_fields:
                        mapped[system_field] = _extract_number(raw_value)
                    elif system_field in date_fields:
                        mapped[system_field] = _extract_date(raw_value)

            workshop = str(mapped.get("workshop", "")).strip()
            product_name = str(mapped.get("product_name", "")).strip()
            batch_no = str(mapped.get("batch_no", "")).strip()
            production_date_str = str(mapped.get("production_date", "")).strip()
            end_date_str = str(mapped.get("end_date", "")).strip()
            weight = float(mapped.get("weight", 0))
            unit = str(mapped.get("unit", "kg")).strip() or "kg"
            notes = str(mapped.get("notes", "")).strip() or None

            if not workshop or not product_name or not batch_no or not production_date_str:
                continue

            try:
                records_data.append(
                    {
                        "workshop": workshop,
                        "product_name": product_name,
                        "batch_no": batch_no,
                        "production_date": date.fromisoformat(production_date_str),
                        "end_date": date.fromisoformat(end_date_str) if end_date_str else None,
                        "weight": weight,
                        "unit": unit,
                        "notes": notes,
                    }
                )
            except ValueError:
                continue

        if not records_data:
            return ApiResponse(code=400, message="未找到有效数据，请检查多维表格字段")

        existing_keys = set()
        for record_data in records_data:
            existing_keys.add(
                f"{record_data['product_name']}|{record_data['workshop']}|"
                f"{record_data['batch_no']}|{record_data['production_date']}"
            )

        # Check existing records in DB using parameterized queries
        existing_in_db: set[str] = set()
        if existing_keys:
            conditions = []
            for k in existing_keys:
                parts = k.split("|")
                conditions.append(
                    and_(
                        ProductOutput.product_name == parts[0],
                        ProductOutput.workshop == parts[1],
                        ProductOutput.batch_no == parts[2],
                        ProductOutput.production_date == parts[3],
                    )
                )
            result = await db.execute(
                select(
                    ProductOutput.product_name,
                    ProductOutput.workshop,
                    ProductOutput.batch_no,
                    ProductOutput.production_date,
                ).where(
                    not_(ProductOutput.is_deleted),
                    or_(*conditions),
                )
            )
            for row in result.fetchall():
                existing_in_db.add(f"{row[0]}|{row[1]}|{row[2]}|{row[3]}")

        filtered_records = []
        skipped_count = 0
        for record_data in records_data:
            batch_key = (
                f"{record_data['product_name']}|{record_data['workshop']}|"
                f"{record_data['batch_no']}|{record_data['production_date']}"
            )
            if batch_key in existing_in_db:
                skipped_count += 1
                continue
            filtered_records.append(record_data)

        if not filtered_records:
            return ApiResponse(
                code=200,
                message=f"所有 {skipped_count} 条记录已存在，无需导入",
                data=ImportResponse(imported=0, skipped=skipped_count, batch_id=""),
            )

        # Match product_id
        product_cache: dict[str, uuid.UUID | None] = {}
        for record_data in filtered_records:
            cache_key = f"{record_data['product_name']}|{record_data['workshop']}"
            if cache_key not in product_cache:
                result = await db.execute(
                    text(
                        "SELECT id FROM production.products WHERE name = :name "
                        "AND workshop = :workshop AND is_deleted = false LIMIT 1"
                    ),
                    {"name": record_data["product_name"], "workshop": record_data["workshop"]},
                )
                result_row = result.first()
                product_cache[cache_key] = result_row[0] if result_row else None
            record_data["product_id"] = product_cache[cache_key]

        service = ProductOutputService(db)
        count = await service.batch_import(filtered_records)
        return ApiResponse(
            data=ImportResponse(imported=count, skipped=skipped_count, batch_id=""),
            message=f"成功导入 {count} 条记录，跳过 {skipped_count} 条重复记录",
        )
    except Exception as e:
        return ApiResponse(code=500, message=f"导入失败：{str(e)}")

"""设备导入 API 路由 (v3 - Smart Mapping & Inference)."""

import base64
import io
import uuid
from datetime import date, datetime
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.deps import CurrentUser
from app.core.response import success_response
from app.modules.equipment import repository as repo
from app.modules.hr.models import HrDepartment

router = APIRouter()

COLUMN_MAPPING = {
    "资产编号": ["资产编号", "编号", "Asset No"],
    "标签号": ["标签号", "Label No"],
    "资产说明": ["资产说明", "设备名称", "Name"],
    "设备位号": ["设备位号", "Tag No"],
    "设备分类": ["设备分类", "Class"],
    "资产类别说明": ["资产类别说明", "类别"],
    "制造商": ["制造商", "厂家"],
    "型号": ["型号", "Model"],
    "设备规格": ["设备规格", "Spec"],
    "供应商": ["供应商", "Vendor"],
    "当前成本": ["当前成本", "金额"],
    "账面净值": ["账面净值", "帐面净值", "净值"],
    "出厂日期": ["出厂日期", "生产日期"],
    "启用日期": ["启用日期", "投用日期"],
    "实物所在部门": ["实物所在部门", "部门", "归属部门"],
    "实物所在地点": ["实物所在地点", "位置", "Location"],
    "负责人": ["负责人", "责任人"],
    "报废状态": ["报废状态", "状态"],
    "报废时间": ["报废时间", "报废日期"],
    "数量": ["数量", "Quantity"],
}

DEPT_MAPPING_V3 = {
    # 职能与班组映射
    "检验室": "质量控制部",
    "质量部": "质量保证部",
    "环保中心": "安全环保部",
    "安全中心": "安全环保部",
    "工程部": "设备工程部",
    "仪表电工班": "设备工程部",
    "机修班": "动力部",
    "制冷班": "动力部",
    "锅炉制水班": "动力部",
    "炊事班": "人事行政部",
    "实验室": "技术研发部",
    "生产管理部": "生产部",

    # 制造部映射
    "头孢精制制造部": "头孢无菌制造部",

    # 车间编号映射 (非头孢)
    "非头孢一车间": "101车间",
    "非头孢二车间": "102车间",
    "非头孢三车间": "103车间",
    "非头孢五车间": "105车间",
    "非头孢六车间": "106车间",
    "非头孢七车间": "107车间",

    # 车间编号映射 (头孢)
    "头孢合成一车间": "201车间",
    "头孢合成二车间": "202车间",
    "头孢精制一车间": "301车间",
    "头孢精制二车间": "302车间",
    "头孢精制三车间": "303车间",

    # 特殊岗位映射 (溶剂回收)
    "溶剂回收车间-401岗": "溶剂回收车间",
    "溶剂回收车间-402岗": "溶剂回收车间",
    "溶剂回收车间-403岗": "溶剂回收车间",
    "溶剂回收车间-404岗": "溶剂回收车间",
    "溶剂回收车间-405岗": "溶剂回收车间",

    # 自映射：确保 Excel 原名与数据库名称完全一致
    "仓库": "仓库",
    "头孢合成制造部": "头孢合成制造部",
}


def get_column_value(row: dict[str, Any], field_name: str) -> Any:
    possible_names = COLUMN_MAPPING.get(field_name, [field_name])
    if not any(name in row for name in possible_names):
        for key in row.keys():
            clean_key = str(key).strip()
            if clean_key in possible_names:
                return row[key]
    for name in possible_names:
        if name in row:
            return row[name]
    return None


async def get_department_id_by_name(db: AsyncSession, dept_name: str) -> uuid.UUID | None:
    from sqlalchemy import select
    result = await db.execute(
        select(HrDepartment.id).where(
            HrDepartment.name == dept_name,
            ~HrDepartment.is_deleted
        )
    )
    return result.scalar_one_or_none()


async def map_department_name_v3(excel_dept: str, db: AsyncSession) -> tuple[str | None, uuid.UUID | None]:
    if not excel_dept:
        return None, None

    # Phase 1: Special handling
    if excel_dept.startswith("溶剂回收车间-"):
        standard_name = "溶剂回收车间"
    else:
        standard_name = DEPT_MAPPING_V3.get(excel_dept, excel_dept)

    # Phase 2 & 3: DB Lookup
    dept_id = await get_department_id_by_name(db, standard_name)
    if dept_id:
        return standard_name, dept_id

    # Fallback
    return None, None


def parse_excel_date(value: Any) -> date | None:
    if not value:
        return None
    # 处理 datetime 对象（openpyxl 读取 Excel 日期单元格时返回）
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, (int, float)):
        from datetime import timedelta
        base_date = date(1899, 12, 30)
        try:
            return base_date + timedelta(days=int(value))
        except (ValueError, OverflowError):
            return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"]:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


# ── v3 Smart Inference Functions ──

def infer_equipment_class(category_description: str | None) -> str:
    if not category_description:
        return "C"
    if "房屋建筑物" in category_description or "房屋" in category_description:
        return "A"
    elif "运输设备" in category_description or "车辆" in category_description:
        return "B"
    elif "电子设备" in category_description or "机器设备" in category_description:
        return "C"
    else:
        return "C"


def infer_importance(current_cost: float | None) -> str:
    if current_cost is None:
        return "中"
    if current_cost > 100000:
        return "高"
    elif current_cost >= 50000:
        return "中"
    else:
        return "低"


def infer_status(scrap_status: str | None) -> str:
    if scrap_status == "未报废":
        return "在用"
    elif scrap_status in ["已报废", "报废"]:
        return "报废"
    else:
        return "在用"


@router.get("/template", summary="下载导入模板")
async def download_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "设备台账"
    headers = list(COLUMN_MAPPING.keys())
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col)].width = 15
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return success_response(data=base64.b64encode(buffer.read()).decode())


@router.post("/preview", summary="预览导入数据")
async def preview_import(data: list[dict[str, Any]], db: AsyncSession = Depends(get_db)):
    results = []
    for idx, row in enumerate(data):
        asset_no = get_column_value(row, "资产编号")
        name = get_column_value(row, "资产说明")
        dept_raw = get_column_value(row, "实物所在部门")
        category_desc = get_column_value(row, "资产类别说明")
        current_cost_raw = get_column_value(row, "当前成本")
        scrap_status = get_column_value(row, "报废状态")
        quantity = get_column_value(row, "数量")

        try:
            current_cost = float(str(current_cost_raw or "0").replace("¥", "").replace(",", ""))
        except (ValueError, TypeError):
            current_cost = None

        equipment_class = infer_equipment_class(category_desc)
        importance = infer_importance(current_cost)
        status = infer_status(scrap_status)

        dept_name, dept_id = await map_department_name_v3(dept_raw, db)

        technical_params = {}
        if quantity is not None:
            technical_params["数量"] = quantity

        errors = []
        warnings = []
        if not asset_no:
            errors.append("资产编号不能为空")
        if not name:
            errors.append("设备名称不能为空")
        if not dept_id and dept_raw:
            warnings.append(f"部门 '{dept_raw}' 未在系统中找到，将设为 NULL")

        results.append({
            "row_index": idx,
            "asset_no": asset_no,
            "name": name,
            "label_no": get_column_value(row, "标签号"),
            "manufacturer": get_column_value(row, "制造商"),
            "model": get_column_value(row, "型号"),
            "location_text": get_column_value(row, "实物所在地点"),
            "department_name": dept_name,
            "department_id": str(dept_id) if dept_id else None,
            "equipment_class": equipment_class,
            "importance": importance,
            "status": status,
            "category_description": category_desc,
            "current_cost": current_cost,
            "technical_params": technical_params,
            "validation_errors": errors,
            "warnings": warnings
        })

    valid_count = sum(1 for r in results if not r["validation_errors"])
    warning_count = sum(1 for r in results if r["warnings"])
    return success_response(data={"total": len(results), "valid_count": valid_count,
        "warning_count": warning_count,
        "items": results})


@router.post("/batch", summary="执行批量导入")
async def batch_import(
    data: list[dict[str, Any]],
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
):
    created = 0
    skipped = 0
    errors = []
    for idx, row in enumerate(data):
        try:
            asset_no = str(get_column_value(row, "资产编号") or "").strip()
            if not asset_no:
                skipped += 1
                continue
            name = str(get_column_value(row, "资产说明") or "")
            if not name:
                skipped += 1
                continue

            dept_raw = get_column_value(row, "实物所在部门")
            dept_name, dept_id = await map_department_name_v3(dept_raw, db)

            try:
                current_cost = float(str(get_column_value(row, "当前成本") or "0").replace("¥", "").replace(",", ""))
            except (ValueError, TypeError):
                current_cost = None

            category_desc = get_column_value(row, "资产类别说明")
            scrap_status = get_column_value(row, "报废状态")
            quantity = get_column_value(row, "数量")

            equipment_class = infer_equipment_class(category_desc)
            importance = infer_importance(current_cost)
            status = infer_status(scrap_status)

            technical_params = {}
            if quantity is not None:
                technical_params["数量"] = quantity

            try:
                book_value = float(str(get_column_value(row, "账面净值") or "0").replace("¥", "").replace(",", ""))
            except (ValueError, TypeError):
                book_value = None

            equipment_data = {
                "asset_no": asset_no, "name": name,
                "label_no": str(get_column_value(row, "标签号") or "") or None,
                "equipment_tag": str(get_column_value(row, "设备位号") or "") or None,
                "equipment_class": equipment_class,
                "category_description": category_desc,
                "manufacturer": str(get_column_value(row, "制造商") or "") or None,
                "model": str(get_column_value(row, "型号") or "") or None,
                "specification": str(get_column_value(row, "设备规格") or "") or None,
                "supplier": str(get_column_value(row, "供应商") or "") or None,
                "current_cost": current_cost,
                "book_value": book_value,
                "production_date": parse_excel_date(get_column_value(row, "出厂日期")),
                "commissioning_date": parse_excel_date(get_column_value(row, "启用日期")),
                "department_id": dept_id,
                "location_text": str(get_column_value(row, "实物所在地点") or "") or None,
                "responsible_person_name": str(get_column_value(row, "负责人") or "") or None,
                "status": status,
                "importance": importance,
                "scrap_status": str(scrap_status or "") or None,
                "scrap_time": parse_excel_date(get_column_value(row, "报废时间")),
                "technical_params": technical_params if technical_params else None,
            }

            existing = await repo.get_equipment_by_asset_no(db, asset_no)
            if existing:
                skipped += 1
                continue
            await repo.create_equipment(db, equipment_data)
            await db.commit()
            created += 1
        except Exception as e:
            await db.rollback()
            skipped += 1
            errors.append({"row": idx, "error": str(e)})
    return success_response(data={"created_count": created, "skipped_count": skipped, "errors": errors})


@router.post("/", summary="上传Excel文件并解析")
async def import_excel(file: UploadFile = File(...)) -> dict[str, Any]:
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 文件")
    from openpyxl import load_workbook
    content_bytes = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content_bytes))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 解析失败: {str(e)}")
    ws = wb.active
    headers = None
    start_row = 2
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] == "资产编号":
            headers = list(row)
            start_row = i + 1
            break
    if not headers:
        headers = [cell.value for cell in ws[1]]
        start_row = 2
    data = []
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if any(cell is not None for cell in row):
            row_dict = {str(headers[i]): row[i] for i in range(len(headers)) if i < len(row) and headers[i]}
            data.append(row_dict)
    return success_response(data=data)

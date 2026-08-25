"""Document text extraction for uploaded attachments.

Supports PDF, DOCX, XLSX, TXT, and Markdown files.
For scanned PDFs (image-only), uses PaddleOCR to extract text.
Uses hybrid approach: PP-OCR for simple text, PP-StructureV3 for structured documents.
"""

import logging
import tempfile
from pathlib import Path

logger = logging.getLogger(__name__)


class DocumentParser:
    """Extract text content from common document formats."""

    SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".xls", ".txt", ".md"}

    @staticmethod
    def extract_text(file_path: str, max_chars: int = 50000) -> str:
        """Extract text from a file. Truncates to *max_chars* to avoid token limits."""
        ext = Path(file_path).suffix.lower()
        if ext == ".pdf":
            text = DocumentParser._extract_pdf(file_path)
        elif ext == ".docx":
            text = DocumentParser._extract_docx(file_path)
        elif ext in (".xlsx", ".xls"):
            text = DocumentParser._extract_xlsx(file_path)
        elif ext in (".txt", ".md"):
            text = DocumentParser._extract_txt(file_path)
        else:
            raise ValueError(f"Unsupported file format: {ext}")

        if len(text) > max_chars:
            text = text[:max_chars] + "\n\n...（文档过长，已截断）"
        return text.strip()

    @staticmethod
    def _extract_pdf(path: str) -> str:
        """Extract text from PDF. Falls back to OCR for scanned PDFs."""
        from pypdf import PdfReader

        reader = PdfReader(path)
        parts: list[str] = []

        # First pass: try to extract text directly
        for page in reader.pages:
            t = page.extract_text()
            if t:
                parts.append(t)

        text = "\n".join(parts)

        # If no text extracted, this is likely a scanned PDF - use OCR
        if len(text.strip()) < 100:  # Less than 100 chars means probably empty
            logger.info(f"PDF appears to be scanned (extracted {len(text)} chars), attempting OCR...")
            text = DocumentParser._extract_pdf_ocr(path, max_pages=10)

        return text

    @staticmethod
    def _extract_pdf_ocr(path: str, max_pages: int = 10) -> str:
        """Extract text from scanned PDF using PaddleOCR (subprocess isolation).

        Uses PP-StructureV3 for better structure preservation (tables, formulas, layout).

        Args:
            path: PDF file path
            max_pages: Maximum pages to process (to avoid timeout)
        """
        try:
            from pypdf import PdfReader, PdfWriter

            from app.shared.ocr_service import OCRService, get_ocr_lock

            # Create a temporary PDF with only the first max_pages pages
            reader = PdfReader(path)
            total_pages = len(reader.pages)
            pages_to_process = min(max_pages, total_pages)

            if pages_to_process == 0:
                return "[PDF 无页面]"

            # Create temporary truncated PDF
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                truncated_path = Path(tmp.name)

            try:
                writer = PdfWriter()
                for i in range(pages_to_process):
                    writer.add_page(reader.pages[i])

                with open(truncated_path, "wb") as f:
                    writer.write(f)

                # Acquire global lock for serialization
                lock = get_ocr_lock()

                with lock:
                    # Call subprocess-based OCR
                    # Timeout: base 120s, formula will use max(120, 120+60*pages)
                    result = OCRService._run_ocr_in_subprocess(
                        file_path=truncated_path,
                        timeout=120,
                        min_memory_gb=8.0,
                    )

                # Extract page texts from result
                pages = result.get("pages", [])
                parts = [
                    f"--- Page {p['page_number']} ---\n{p.get('markdown', '').strip()}"
                    for p in pages
                    if p.get("markdown", "").strip()
                ]

                if not parts:
                    return "[OCR未能提取到文本内容]"

                return "\n\n".join(parts)

            finally:
                # Clean up temporary PDF
                if truncated_path.exists():
                    truncated_path.unlink()

        except RuntimeError as e:
            error_msg = str(e)
            logger.error(f"OCR extraction failed: {error_msg}")
            return f"[OCR提取失败: {error_msg}]"
        except Exception as e:
            logger.error(f"OCR extraction failed: {e}")
            return f"[OCR提取失败: {str(e)}]"

    @staticmethod
    def _extract_docx(path: str) -> str:
        from docx import Document

        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs if p.text)

    @staticmethod
    def _extract_xlsx(path: str) -> str:
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True)
        parts: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"## Sheet: {sheet_name}")
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                row_str = "\t".join(str(cell) if cell is not None else "" for cell in row)
                if row_str.strip():
                    rows.append(row_str)
            parts.append("\n".join(rows))
        wb.close()
        return "\n\n".join(parts)

    @staticmethod
    def _extract_txt(path: str) -> str:
        with open(path, encoding="utf-8", errors="replace") as f:
            return f.read()
